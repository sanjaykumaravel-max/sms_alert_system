import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

class ExcelManager:
    """Manages Excel file operations for the SMS alert app."""
    
    def __init__(self, filepath):
        """Initialize Excel manager with file path."""
        self.filepath = Path(filepath)
    
    def create_workbook(self):
        """Create a new Excel workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SMS Alerts"
        return wb, ws
    
    def add_headers(self, ws, headers):
        """Add header row to worksheet."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    def add_row(self, ws, data):
        """Add a data row to worksheet."""
        ws.append(data)
    
    def save(self, wb):
        """Save workbook to file."""
        self.filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.filepath)
    
    def read_data(self):
        """Read data from Excel file."""
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        return ws.values
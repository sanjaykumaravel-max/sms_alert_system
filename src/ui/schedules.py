"""
Professional desktop UI for managing SMS schedules.

Features:
- List schedules
- Add / Edit / Delete / Enable / Disable
- Simple recurrence: one-time or daily/weekly
- Save/load schedules to JSON in project data folder
- Validation and next-run preview

Dependencies: Python 3.8+, standard library (tkinter)
"""
from __future__ import annotations
import json
import uuid
from pathlib import Path
from dataclasses import dataclass, asdict, field
from datetime import datetime, time, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from . import theme as theme_mod
from typing import List, Optional
import pandas as pd
try:
    from ..app_paths import data_dir, exports_dir
except Exception:
    from app_paths import data_dir, exports_dir
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except Exception:
    load_workbook = None
    Font = PatternFill = Alignment = Border = Side = None
import customtkinter as ctk
from .validation import normalize_phone_input, validate_phone

DATA_DIR = data_dir()
SCHEDULES_FILE = DATA_DIR / "schedules.json"
DATE_FORMAT = "%Y-%m-%d %H:%M"


@dataclass
class Schedule:
    id: str
    name: str
    phone: str
    message: str
    kind: str  # "one-time", "daily", "weekly"
    run_at: str  # one-time: "YYYY-MM-DD HH:MM", recurring: "HH:MM" or "HH:MM:SS"
    weekdays: List[int] = field(default_factory=list)  # 0=Mon .. 6=Sun (for weekly)
    enabled: bool = True
    created_at: Optional[str] = None

    def next_run(self, from_dt: Optional[datetime] = None) -> Optional[datetime]:
        now = from_dt or datetime.now()
        if not self.enabled:
            return None

        if self.kind == "one-time":
            try:
                dt = datetime.strptime(self.run_at, DATE_FORMAT)
            except Exception:
                return None
            return dt if dt >= now else None

        # recurring (daily/weekly)
        try:
            t = time.fromisoformat(self.run_at)
        except Exception:
            return None

        # candidate today at time t
        candidate = datetime.combine(now.date(), t)
        if candidate >= now:
            start = candidate
        else:
            start = candidate + timedelta(days=1)

        if self.kind == "daily":
            return start

        if self.kind == "weekly":
            # search up to 14 days for next matching weekday
            for offset in range(0, 14):
                dt = start + timedelta(days=offset)
                if dt.weekday() in self.weekdays:
                    return dt.replace(hour=t.hour, minute=t.minute, second=0, microsecond=0)
        return None


def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not SCHEDULES_FILE.exists():
        SCHEDULES_FILE.write_text("[]", encoding="utf-8")


def load_schedules() -> List[Schedule]:
    ensure_data_dir()
    try:
        raw = SCHEDULES_FILE.read_text(encoding="utf-8")
        data = json.loads(raw or "[]")
        return [Schedule(**item) for item in data]
    except Exception:
        return []


def save_schedules(schedules: List[Schedule]):
    ensure_data_dir()
    data = [asdict(s) for s in schedules]
    SCHEDULES_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


class ScheduleDialog(simpledialog.Dialog):
    def __init__(self, parent, title, schedule: Optional[Schedule] = None):
        self.schedule = schedule
        super().__init__(parent, title)

    def body(self, master):
        self.resizable(False, False)
        label_color = "#e2e8f0"
        field_bg = "#0b1220"
        try:
            master.configure(bg=theme_mod.SIMPLE_PALETTE.get("card", "#0f1724"))
        except Exception:
            pass

        ctk.CTkLabel(master, text="Name:", text_color=label_color, font=("Segoe UI", 13)).grid(row=0, column=0, sticky="w")
        self.name_var = tk.StringVar(value=getattr(self.schedule, "name", ""))
        ctk.CTkEntry(master, textvariable=self.name_var, width=250, font=("Segoe UI", 13), height=34).grid(row=0, column=1, columnspan=3, pady=2)

        ctk.CTkLabel(master, text="Phone:", text_color=label_color, font=("Segoe UI", 13)).grid(row=1, column=0, sticky="w")
        self.phone_var = tk.StringVar(value=getattr(self.schedule, "phone", ""))
        ctk.CTkEntry(master, textvariable=self.phone_var, width=200, font=("Segoe UI", 13), height=34).grid(row=1, column=1, pady=2)

        ctk.CTkLabel(master, text="Message:", text_color=label_color, font=("Segoe UI", 13)).grid(row=2, column=0, sticky="nw")
        self.message_text = tk.Text(
            master,
            width=50,
            height=5,
            font=("Segoe UI", 11),
            bg=field_bg,
            fg=label_color,
            insertbackground=label_color,
            relief="flat",
            highlightthickness=1,
            highlightbackground="#1f2937",
            highlightcolor="#1f2937",
        )
        self.message_text.grid(row=2, column=1, columnspan=3, pady=2)
        if self.schedule:
            self.message_text.insert("1.0", self.schedule.message)

        ctk.CTkLabel(master, text="Type:", text_color=label_color, font=("Segoe UI", 13)).grid(row=3, column=0, sticky="w")
        self.kind_var = tk.StringVar(value=getattr(self.schedule, "kind", "one-time"))
        ctk.CTkOptionMenu(master, variable=self.kind_var, values=["one-time", "daily", "weekly"], width=150, font=("Segoe UI", 12)).grid(row=3, column=1, pady=2)

        ctk.CTkLabel(master, text="Run (one-time: YYYY-MM-DD HH:MM, recurring: HH:MM):", text_color="#cbd5e1", font=("Segoe UI", 11)).grid(row=4, column=0, columnspan=2, sticky="w")
        self.run_at_var = tk.StringVar(value=getattr(self.schedule, "run_at", datetime.now().strftime(DATE_FORMAT)))
        ctk.CTkEntry(master, textvariable=self.run_at_var, width=200, font=("Segoe UI", 13), height=34).grid(row=4, column=2, pady=2)

        ctk.CTkLabel(master, text="Weekdays (Mon=0..Sun=6, comma sep):", text_color="#cbd5e1", font=("Segoe UI", 11)).grid(row=5, column=0, columnspan=2, sticky="w")
        self.weekdays_var = tk.StringVar(value=",".join(map(str, getattr(self.schedule, "weekdays", []))))
        ctk.CTkEntry(master, textvariable=self.weekdays_var, width=200, font=("Segoe UI", 13), height=34).grid(row=5, column=2, pady=2)

        self.enabled_var = tk.BooleanVar(value=getattr(self.schedule, "enabled", True))
        ctk.CTkCheckBox(master, text="Enabled", variable=self.enabled_var, text_color=label_color, font=("Segoe UI", 12)).grid(row=6, column=0, sticky="w", pady=4)

        return master

    def validate(self):
        name = self.name_var.get().strip()
        phone = self.phone_var.get().strip()
        message = self.message_text.get("1.0", "end").strip()
        kind = self.kind_var.get()
        run_at = self.run_at_var.get().strip()

        if not name:
            messagebox.showerror("Validation", "Name is required.")
            return False
        if not phone:
            messagebox.showerror("Validation", "Phone is required.")
            return False
        if not validate_phone(phone, "Phone"):
            return False
        if not message:
            messagebox.showerror("Validation", "Message is required.")
            return False

        if kind == "one-time":
            try:
                datetime.strptime(run_at, DATE_FORMAT)
            except Exception:
                messagebox.showerror("Validation", f"One-time runs must use format: {DATE_FORMAT}")
                return False
        else:
            try:
                time.fromisoformat(run_at)
            except Exception:
                messagebox.showerror("Validation", "Recurring time must be HH:MM or HH:MM:SS")
                return False

        if kind == "weekly":
            try:
                if self.weekdays_var.get().strip() == "":
                    messagebox.showerror("Validation", "Weekly schedules require at least one weekday (0..6).")
                    return False
                weekdays = [int(x) for x in self.weekdays_var.get().split(",") if x.strip() != ""]
                for d in weekdays:
                    if d < 0 or d > 6:
                        raise ValueError()
            except Exception:
                messagebox.showerror("Validation", "Weekdays must be comma-separated integers 0..6.")
                return False

        return True

    def apply(self):
        name = self.name_var.get().strip()
        phone = self.phone_var.get().strip()
        message = self.message_text.get("1.0", "end").strip()
        kind = self.kind_var.get()
        run_at = self.run_at_var.get().strip()
        weekdays = [int(x) for x in self.weekdays_var.get().split(",") if x.strip() != ""] if kind == "weekly" else []
        enabled = self.enabled_var.get()
        now_iso = datetime.now().isoformat()
        if self.schedule:
            sid = self.schedule.id
            created = self.schedule.created_at
        else:
            sid = str(uuid.uuid4())
            created = now_iso

        # normalize recurring run_at to include seconds if missing
        if kind != "one-time" and ":" in run_at and run_at.count(":") == 1:
            run_at = f"{run_at}:00"

        self.result = Schedule(
            id=sid,
            name=name,
            phone=normalize_phone_input(phone) or phone,
            message=message,
            kind=kind,
            run_at=run_at,
            weekdays=weekdays,
            enabled=enabled,
            created_at=created,
        )


class SchedulesFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        self.parent = parent
        self.schedules: List[Schedule] = []
        self._surface = theme_mod.SIMPLE_PALETTE.get("card", "#0f1724")
        self._surface_alt = "#0b1220"
        self._surface_header = "#1f2937"
        self._text_primary = "#e2e8f0"
        self._text_muted = "#94a3b8"
        self._accent = theme_mod.SIMPLE_PALETTE.get("primary", "#06B6D4")
        self._build_ui()
        self.load()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=14)
        header.pack(fill="x", pady=(10, 8), padx=10)
        ctk.CTkLabel(header, text="Schedules", font=("Segoe UI Semibold", 20), text_color=self._text_primary).pack(anchor="w", padx=12, pady=(10, 2))
        ctk.CTkLabel(header, text="Plan recurring and one-time SMS alerts", font=("Segoe UI", 12), text_color=self._text_muted).pack(anchor="w", padx=12, pady=(0, 10))

        toolbar = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=12)
        toolbar.pack(fill="x", pady=(0, 8), padx=10)

        btn_font = ("Segoe UI Semibold", 13)
        ctk.CTkButton(toolbar, text="Add", command=self.add_schedule, height=34, font=btn_font, fg_color=self._accent, hover_color="#0891b2").pack(side="left", padx=6, pady=8)
        ctk.CTkButton(toolbar, text="Edit", command=self.edit_schedule, height=34, font=btn_font, fg_color="#334155", hover_color="#475569").pack(side="left", padx=(0, 6), pady=8)
        ctk.CTkButton(toolbar, text="Delete", command=self.delete_schedule, height=34, font=btn_font, fg_color="#b91c1c", hover_color="#991b1b").pack(side="left", padx=(0, 6), pady=8)
        ctk.CTkButton(toolbar, text="Enable/Disable", command=self.toggle_enabled, height=34, font=btn_font, fg_color="#0f766e", hover_color="#115e59").pack(side="left", padx=(0, 6), pady=8)
        ctk.CTkButton(toolbar, text="Refresh", command=self.load, height=34, font=btn_font, fg_color="#334155", hover_color="#475569").pack(side="right", padx=(0, 6), pady=8)
        ctk.CTkButton(toolbar, text="Export", command=self._export_schedules, height=34, font=btn_font, fg_color="#2563eb", hover_color="#1d4ed8").pack(side="right", padx=(0, 6), pady=8)

        content = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=14)
        content.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.tree = ttk.Treeview(content, columns=("name", "phone", "type", "next", "enabled"), show="headings", selectmode="browse", style="Schedules.Treeview")
        try:
            style = ttk.Style()
            style.configure(
                "Schedules.Treeview",
                font=("Segoe UI", 12),
                rowheight=36,
                background=self._surface_alt,
                fieldbackground=self._surface_alt,
                foreground=self._text_primary,
                borderwidth=0,
                relief="flat",
            )
            style.configure(
                "Schedules.Treeview.Heading",
                font=("Segoe UI Semibold", 12),
                background=self._surface_header,
                foreground=self._text_primary,
                relief="flat",
            )
            style.map(
                "Schedules.Treeview",
                background=[("selected", self._accent)],
                foreground=[("selected", "#ffffff")],
            )
        except Exception:
            pass
        try:
            self.tree.tag_configure("even", background=self._surface_alt, foreground=self._text_primary)
            self.tree.tag_configure("odd", background="#111827", foreground=self._text_primary)
        except Exception:
            pass
        self.tree.heading("name", text="Name")
        self.tree.heading("phone", text="Phone")
        self.tree.heading("type", text="Type")
        self.tree.heading("next", text="Next Run")
        self.tree.heading("enabled", text="Enabled")
        self.tree.column("name", width=180)
        self.tree.column("phone", width=120)
        self.tree.column("type", width=80)
        self.tree.column("next", width=180)
        self.tree.column("enabled", width=70, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree.bind("<Double-1>", lambda e: self.edit_schedule())

        status = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=12)
        status.pack(fill="x", pady=(8, 10), padx=10)
        self.count_label = ctk.CTkLabel(status, text="0 schedules", font=("Segoe UI", 13), text_color=self._text_muted)
        self.count_label.pack(side="left", padx=12, pady=8)

    def load(self):
        self.schedules = load_schedules()
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for idx, s in enumerate(sorted(self.schedules, key=lambda x: x.name.lower())):
            nr = s.next_run()
            nr_str = nr.strftime(DATE_FORMAT) if nr else "-"
            tag = "even" if (idx % 2 == 0) else "odd"
            self.tree.insert("", "end", iid=s.id, values=(s.name, s.phone, s.kind, nr_str, "Yes" if s.enabled else "No"), tags=(tag,))
        self.count_label.configure(text=f"{len(self.schedules)} schedules")

    def _export_schedules(self) -> None:
        """Export schedules to CSV and styled Excel."""
        try:
            export_dir = exports_dir()
            export_dir.mkdir(parents=True, exist_ok=True)
            ts = __import__('datetime').datetime.now().strftime("%Y%m%d_%H%M%S")
            fname_csv = export_dir / f"schedules_export_{ts}.csv"
            fname_xlsx = export_dir / f"schedules_export_{ts}.xlsx"
            rows = [asdict(s) for s in self.schedules]
            df = pd.DataFrame(rows)
            df.to_csv(fname_csv, index=False)
            try:
                df.to_excel(fname_xlsx, index=False)
                if load_workbook and Font is not None:
                    wb = load_workbook(str(fname_xlsx))
                    ws = wb.active
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="4F81BD")
                    for cell in list(ws[1]):
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    thin = Side(border_style="thin", color="CCCCCC")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        rnum = row[0].row
                        fill = PatternFill("solid", fgColor=("F7F7F7" if rnum % 2 == 0 else "FFFFFF"))
                        for cell in row:
                            cell.border = border
                            cell.fill = fill
                            cell.alignment = Alignment(vertical='top')
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                val = str(cell.value) if cell.value is not None else ""
                            except Exception:
                                val = ""
                            if len(val) > max_length:
                                max_length = len(val)
                        ws.column_dimensions[col_letter].width = max_length + 2
                    wb.save(str(fname_xlsx))
            except Exception:
                import logging
                logging.getLogger(__name__).exception("Failed to write styled Excel for schedules; CSV saved")
            try:
                messagebox.showinfo("Exported", f"Exported schedules to:\n{fname_csv}\n{fname_xlsx}")
            except Exception:
                pass
        except Exception as exc:
            import logging
            logging.getLogger(__name__).exception("Failed to export schedules: %s", exc)
            try:
                messagebox.showerror("Export error", f"Failed to export schedules: {exc}")
            except Exception:
                pass

    def _get_selected(self) -> Optional[Schedule]:
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Please select a schedule first.")
            return None
        sid = sel[0]
        for s in self.schedules:
            if s.id == sid:
                return s
        return None

    def add_schedule(self):
        dlg = ScheduleDialog(self.winfo_toplevel(), "Add Schedule")
        if getattr(dlg, "result", None):
            self.schedules.append(dlg.result)
            save_schedules(self.schedules)
            self._refresh_tree()

    def edit_schedule(self):
        s = self._get_selected()
        if not s:
            return
        dlg = ScheduleDialog(self.winfo_toplevel(), "Edit Schedule", schedule=s)
        if getattr(dlg, "result", None):
            # replace existing
            self.schedules = [dlg.result if x.id == s.id else x for x in self.schedules]
            save_schedules(self.schedules)
            self._refresh_tree()

    def delete_schedule(self):
        s = self._get_selected()
        if not s:
            return
        if messagebox.askyesno("Delete", f"Delete schedule '{s.name}'?"):
            self.schedules = [x for x in self.schedules if x.id != s.id]
            save_schedules(self.schedules)
            self._refresh_tree()

    def toggle_enabled(self):
        s = self._get_selected()
        if not s:
            return
        for x in self.schedules:
            if x.id == s.id:
                x.enabled = not x.enabled
                break
        save_schedules(self.schedules)
        self._refresh_tree()


if __name__ == "__main__":
    root = tk.Tk()
    root.title("SMS Alert - Schedules")
    root.geometry("780x480")
    frame = SchedulesFrame(root)
    frame.pack(fill="both", expand=True)
    root.mainloop()

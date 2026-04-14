import customtkinter as ctk
import tkinter as tk
import json
import threading
import pandas as pd
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except Exception:
    load_workbook = None
    Font = PatternFill = Alignment = Border = Side = None

from .cards import GIFSpinner
from sms_service import default_sms_service
from api_client import sync_get_operators, sync_create_operator
from . import theme as theme_mod
from .gradient import GradientPanel
from authz import has_role
from .validation import normalize_phone_input, validate_phone, validate_required
try:
    from ..app_paths import data_dir
except Exception:
    from app_paths import data_dir

DATA_DIR = data_dir()
OPERATORS_FILE = DATA_DIR / "operators.json"


def load_operators():
    """Load operators from API or fallback to file."""
    try:
        return sync_get_operators()
    except Exception as e:
        print(f"API not available, falling back to file: {e}")
        if not OPERATORS_FILE.exists():
            return []
        with open(OPERATORS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)


def save_operators(ops):
    """Save operators - mainly for backward compatibility."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(OPERATORS_FILE, "w", encoding="utf-8") as f:
        json.dump(ops, f, indent=2)


class OperatorsFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        self._surface = theme_mod.SIMPLE_PALETTE.get("card", "#0f1724")
        self._surface_alt = "#0b1220"
        self._text_primary = "#e2e8f0"
        self._text_muted = "#94a3b8"
        self._accent = theme_mod.SIMPLE_PALETTE.get("primary", "#06B6D4")

        header = GradientPanel(
            self,
            colors=getattr(theme_mod, "SECTION_GRADIENTS", {}).get("operators", ("#081919", "#0f766e", "#2dd4bf")),
            corner_radius=16,
            border_color="#133b38",
        )
        header.pack(fill="x", padx=12, pady=(12, 8))
        ctk.CTkLabel(
            header.content,
            text="Operators",
            font=("Segoe UI Semibold", 20),
            text_color=self._text_primary,
        ).pack(anchor="w", padx=12, pady=(10, 2))
        ctk.CTkLabel(
            header.content,
            text="Manage contacts and trigger notifications",
            font=("Segoe UI", 13),
            text_color="#ccfbf1",
        ).pack(anchor="w", padx=12, pady=(0, 10))

        list_card = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=14)
        list_card.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        self.listbox = tk.Listbox(
            list_card,
            height=10,
            font=("Segoe UI", 13),
            bg=self._surface_alt,
            fg=self._text_primary,
            selectbackground=self._accent,
            selectforeground="#ffffff",
            relief="flat",
            borderwidth=0,
            highlightthickness=0,
        )
        self.listbox.pack(fill="both", expand=True, pady=10, padx=10)

        # Loading spinner while operators load
        self._spinner = GIFSpinner(self)
        self._spinner.place(relx=0.5, rely=0.3, anchor="center")
        try:
            self._spinner.start()
        except Exception:
            pass

        form = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=12)
        form.pack(fill="x", padx=12, pady=(0, 8))
        self.name_entry = ctk.CTkEntry(form, placeholder_text="Name", font=("Segoe UI", 13), height=36)
        self.name_entry.grid(row=0, column=0, padx=(10, 8), pady=10, sticky="ew")
        self.phone_entry = ctk.CTkEntry(form, placeholder_text="Phone (+countrycode)", font=("Segoe UI", 13), height=36)
        self.phone_entry.grid(row=0, column=1, padx=(0, 10), pady=10, sticky="ew")
        form.grid_columnconfigure(0, weight=1)
        form.grid_columnconfigure(1, weight=1)

        btn_frame = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=12)
        btn_frame.pack(fill="x", padx=12, pady=(0, 8))
        btn_font = ("Segoe UI Semibold", 13)

        add_btn = ctk.CTkButton(
            btn_frame,
            text="Add",
            command=self.add_operator,
            font=btn_font,
            height=34,
            fg_color=self._accent,
            hover_color="#0891b2",
        )
        add_btn.pack(side="left", padx=6, pady=8)

        remove_btn = ctk.CTkButton(
            btn_frame,
            text="Remove Selected",
            command=self.remove_selected,
            font=btn_font,
            height=34,
            fg_color="#b91c1c",
            hover_color="#991b1b",
        )
        remove_btn.pack(side="left", padx=6, pady=8)

        send_sms_btn = ctk.CTkButton(
            btn_frame,
            text="Send SMS to Selected",
            command=self.send_sms_to_selected,
            font=btn_font,
            height=34,
            fg_color="#0f766e",
            hover_color="#115e59",
        )
        send_sms_btn.pack(side="left", padx=6, pady=8)

        save_btn = ctk.CTkButton(
            btn_frame,
            text="Save",
            command=self.save,
            font=btn_font,
            height=34,
            fg_color="#059669",
            hover_color="#047857",
        )
        save_btn.pack(side="right", padx=6, pady=8)

        export_btn = ctk.CTkButton(
            btn_frame,
            text="Export",
            command=self._export_operators,
            font=btn_font,
            height=34,
            fg_color="#2563eb",
            hover_color="#1d4ed8",
        )
        export_btn.pack(side="right", padx=6, pady=8)

        self.status_label = ctk.CTkLabel(
            self,
            text="Loading operators...",
            font=("Segoe UI", 13),
            text_color=self._text_muted,
        )
        self.status_label.pack(anchor="w", padx=14, pady=(0, 6))

        self.operators = []

        def _load_bg():
            try:
                ops = load_operators()
            except Exception:
                ops = []

            def _on_done():
                self.operators = ops or []
                try:
                    self._spinner.stop()
                except Exception:
                    pass
                try:
                    self._spinner.place_forget()
                except Exception:
                    pass
                self.refresh_list()
                try:
                    self.status_label.configure(text=f"{len(self.operators)} operators loaded")
                except Exception:
                    pass

            try:
                self.after(10, _on_done)
            except Exception:
                _on_done()

        threading.Thread(target=_load_bg, daemon=True).start()

    def refresh_list(self):
        self.listbox.delete(0, tk.END)
        for op in self.operators:
            self.listbox.insert(tk.END, f"{op.get('name', '')} - {op.get('phone', '')}")

    def add_operator(self):
        name = self.name_entry.get().strip()
        phone = self.phone_entry.get().strip()
        if not validate_required(name, "Name"):
            return
        if not validate_phone(phone, "Phone"):
            return
        normalized_phone = normalize_phone_input(phone) or phone

        try:
            new_operator = sync_create_operator({"name": name, "phone": normalized_phone, "active": True})
            self.operators.append(new_operator)
            self.name_entry.delete(0, tk.END)
            self.phone_entry.delete(0, tk.END)
            self.refresh_list()
        except Exception as e:
            print(f"API not available, falling back to local addition: {e}")
            self.operators.append({"name": name, "phone": normalized_phone})
            self.name_entry.delete(0, tk.END)
            self.phone_entry.delete(0, tk.END)
            self.refresh_list()
        try:
            self.status_label.configure(text=f"{len(self.operators)} operators loaded")
        except Exception:
            pass

    def remove_selected(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        try:
            user = getattr(self, "dashboard", None) and getattr(self.dashboard, "user", None)
            if not user or not has_role(user, "admin"):
                tk.messagebox.showerror("Permission denied", "Only administrators may remove operators.")
                return
        except Exception:
            tk.messagebox.showerror("Permission denied", "Unable to verify permissions.")
            return
        idx = sel[0]
        self.operators.pop(idx)
        self.refresh_list()
        try:
            self.status_label.configure(text=f"{len(self.operators)} operators loaded")
        except Exception:
            pass

    def save(self):
        save_operators(self.operators)

    def send_sms_to_selected(self):
        sel = self.listbox.curselection()
        if not sel:
            tk.messagebox.showwarning("No Selection", "Please select an operator to send SMS.")
            return
        idx = sel[0]
        op = self.operators[idx]
        phone = op.get("phone")
        if not phone or not validate_phone(str(phone), "Operator Phone"):
            tk.messagebox.showerror("No Phone", "Selected operator has no phone number.")
            return
        phone = normalize_phone_input(str(phone)) or str(phone)
        try:
            user = getattr(self, "dashboard", None) and getattr(self.dashboard, "user", None)
            allowed = bool(user and (has_role(user, "admin") or has_role(user, "sms")))
            if not allowed:
                tk.messagebox.showerror("Permission denied", "You are not allowed to send SMS from this UI.")
                return
        except Exception:
            tk.messagebox.showerror("Permission denied", "Unable to verify permissions.")
            return

        message = f"Hello {op.get('name', 'Operator')}, this is a test SMS from the system."
        try:
            try:
                tk.messagebox.showinfo("SMS", f"Sending SMS to {op.get('name')} at {phone}...")
            except Exception:
                pass

            def _cb(result):
                def _ui():
                    try:
                        if result.get("success"):
                            tk.messagebox.showinfo("SMS Sent", f"SMS successfully sent to {op.get('name')} at {phone}")
                        else:
                            err = result.get("error") or result.get("response_text") or "Unknown error"
                            detail = result.get("response_text")
                            if detail:
                                tk.messagebox.showerror(
                                    "SMS Failed",
                                    f"Failed to send SMS to {op.get('name')}: {err}\n\nProvider response: {detail}",
                                )
                                return
                            tk.messagebox.showerror("SMS Failed", f"Failed to send SMS to {op.get('name')}: {err}")
                    except Exception:
                        pass

                try:
                    self.after(10, _ui)
                except Exception:
                    _ui()

            default_sms_service.send_async(phone, message, callback=_cb)
        except Exception as e:
            tk.messagebox.showerror("SMS Error", f"Failed to send SMS: {str(e)}")

    def _sms_callback(self, result):
        if result.get("success"):
            print("SMS sent successfully")
        else:
            print(f"SMS failed: {result.get('error')}")

    def _export_operators(self) -> None:
        """Export operators list to CSV and styled Excel (if openpyxl available)."""
        try:
            export_dir = DATA_DIR / "exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            ts = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
            fname_csv = export_dir / f"operators_export_{ts}.csv"
            fname_xlsx = export_dir / f"operators_export_{ts}.xlsx"
            df = pd.DataFrame(self.operators or [])
            df.to_csv(fname_csv, index=False)
            try:
                df.to_excel(fname_xlsx, index=False)
                if load_workbook and Font is not None:
                    wb = load_workbook(str(fname_xlsx))
                    ws = wb.active
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="4F81BD")
                    for cell in list(ws[1]):
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    thin = Side(border_style="thin", color="CCCCCC")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        rnum = row[0].row
                        fill = PatternFill("solid", fgColor=("F7F7F7" if rnum % 2 == 0 else "FFFFFF"))
                        for cell in row:
                            cell.border = border
                            cell.fill = fill
                            cell.alignment = Alignment(vertical="top")
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                val = str(cell.value) if cell.value is not None else ""
                            except Exception:
                                val = ""
                            if len(val) > max_length:
                                max_length = len(val)
                        ws.column_dimensions[col_letter].width = max_length + 2
                    wb.save(str(fname_xlsx))
            except Exception:
                LOG = __import__("logging").getLogger(__name__)
                LOG.exception("Failed to write styled Excel for operators; CSV saved")
            try:
                tk.messagebox.showinfo("Exported", f"Exported operators to:\n{fname_csv}\n{fname_xlsx}")
            except Exception:
                pass
        except Exception as exc:
            LOG = __import__("logging").getLogger(__name__)
            LOG.exception("Failed to export operators: %s", exc)
            try:
                tk.messagebox.showerror("Export error", f"Failed to export operators: {exc}")
            except Exception:
                pass

import os
import datetime
import logging
from html import escape
from typing import Optional

import pandas as pd
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, ttk
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except Exception:
    load_workbook = None
    Font = PatternFill = Alignment = Border = Side = None
from . import theme as theme_mod
from .gradient import GradientPanel
from authz import has_role
from .dialogs import create_dialog
from .validation import (
    normalize_date_input,
    normalize_phone_input,
    validate_date_string,
    validate_number,
    validate_optional_phone,
    validate_required,
)
try:
    from ..app_paths import data_dir, exports_dir
    from ..machine_store import (
        complete_machine_maintenance,
        DEFAULT_COLUMNS,
        archive_machine,
        evaluate_machine_status,
        machine_history,
        load_machines,
        save_machines,
    )
    from ..mine_store import get_active_mine
except Exception:
    from app_paths import data_dir, exports_dir
    from machine_store import (
        complete_machine_maintenance,
        DEFAULT_COLUMNS,
        archive_machine,
        evaluate_machine_status,
        machine_history,
        load_machines,
        save_machines,
    )
    from mine_store import get_active_mine

LOG = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

DATA_DIR = str(data_dir())


class MachinesWindow(ctk.CTkToplevel):
    """Top-level window that hosts a MachinesFrame."""

    def __init__(self, parent: ctk.CTkBaseClass) -> None:
        super().__init__(parent)
        self.title("Machines")
        self.geometry("900x560")
        self.parent = parent

        self.frame = MachinesFrame(self)
        self.frame.pack(fill="both", expand=True)


class MachinesFrame(ctk.CTkFrame):
    """Frame that shows machines in a table with a details pane."""

    def __init__(self, parent: ctk.CTkBaseClass, dashboard: Optional[object] = None) -> None:
        super().__init__(parent, fg_color="transparent")
        self.parent = parent
        self.dashboard = dashboard

        self.df = pd.DataFrame(columns=DEFAULT_COLUMNS)
        self._load_machines()

        # Modern visual tokens for a richer workspace look
        self._surface = theme_mod.SIMPLE_PALETTE.get("card", "#111827")
        self._surface_alt = "#0b1220"
        self._surface_header = "#1f2937"
        self._text_primary = "#e2e8f0"
        self._text_muted = "#94a3b8"
        self._accent = theme_mod.SIMPLE_PALETTE.get("primary", "#2563eb")

        # layout: left = table, right = details
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)

        header = GradientPanel(
            self,
            colors=getattr(theme_mod, "SECTION_GRADIENTS", {}).get("machines", ("#081a22", "#0369a1", "#38bdf8")),
            corner_radius=16,
            border_color="#1d2a3f",
        )
        header.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 6))
        ctk.CTkLabel(
            header.content,
            text="Machines Workspace",
            font=("Segoe UI Semibold", 24),
            text_color="#f8fafc",
        ).pack(anchor="w", padx=16, pady=(14, 4))
        ctk.CTkLabel(
            header.content,
            text="Manage the machine register, service thresholds, runtime-based maintenance, and completion workflow from one place.",
            font=("Segoe UI", 13),
            text_color="#dbeafe",
            wraplength=980,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 14))

        left = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=14)
        left.grid(row=1, column=0, sticky="nsew", padx=8, pady=8)

        right = ctk.CTkFrame(self, width=380, fg_color=self._surface, corner_radius=14)
        # place details pane but keep it hidden so the table initially fills the area
        right.grid(row=1, column=1, sticky="ns", padx=8, pady=8)
        right.grid_remove()
        self._right = right
        self._right_visible = False

        # Toolbar
        toolbar = ctk.CTkFrame(self, fg_color=self._surface, corner_radius=12)
        toolbar.grid(row=2, column=0, columnspan=2, sticky="ew", padx=8, pady=(0,8))
        toolbar.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6, 7), weight=1)

        btn_font = ("Segoe UI Semibold", 13)
        self.add_btn = ctk.CTkButton(
            toolbar,
            text="Add",
            command=self._open_add_dialog,
            font=btn_font,
            height=34,
            fg_color=self._accent,
            hover_color="#1d4ed8",
        )
        self.add_btn.grid(row=0, column=0, padx=4, pady=6, sticky="w")
        self.edit_btn = ctk.CTkButton(
            toolbar,
            text="Edit",
            command=self._edit_selected,
            font=btn_font,
            height=34,
            fg_color="#334155",
            hover_color="#475569",
        )
        self.edit_btn.grid(row=0, column=1, padx=4, pady=6)
        self.delete_btn = ctk.CTkButton(
            toolbar,
            text="Delete",
            fg_color="#b91c1c",
            hover_color="#991b1b",
            command=self._delete_selected,
            font=btn_font,
            height=34,
        )
        self.delete_btn.grid(row=0, column=2, padx=4, pady=6)
        self.save_btn = ctk.CTkButton(
            toolbar,
            text="Save",
            command=self._save_machines,
            font=btn_font,
            height=34,
            fg_color="#059669",
            hover_color="#047857",
        )
        self.save_btn.grid(row=0, column=3, padx=4, pady=6)
        self.complete_btn = ctk.CTkButton(
            toolbar,
            text="Complete Maint",
            command=self._complete_selected_maintenance,
            font=btn_font,
            height=34,
            fg_color="#7c3aed",
            hover_color="#6d28d9",
        )
        self.complete_btn.grid(row=0, column=4, padx=4, pady=6)
        self.report_btn = ctk.CTkButton(
            toolbar,
            text="Print Report",
            command=self._print_completion_report_selected,
            font=btn_font,
            height=34,
            fg_color="#1d4ed8",
            hover_color="#1e40af",
        )
        self.report_btn.grid(row=0, column=5, padx=4, pady=6)
        self.export_btn = ctk.CTkButton(
            toolbar,
            text="Export CSV",
            command=self._export_csv,
            font=btn_font,
            height=34,
            fg_color="#0f766e",
            hover_color="#115e59",
        )
        self.export_btn.grid(row=0, column=6, padx=4, pady=6)
        # Apply UI-level permissions: disable/hide buttons for unauthorized users
        try:
            user = getattr(self, 'dashboard', None) and getattr(self.dashboard, 'user', None)
            can_manage = bool(user and (has_role(user, 'admin') or has_role(user, 'maintainer')))
            can_export = bool(user and (has_role(user, 'admin') or has_role(user, 'reporter')))
            if not can_manage:
                try:
                    self.add_btn.configure(state='disabled')
                except Exception:
                    try:
                        self.add_btn.configure(fg_color='gray')
                    except Exception:
                        pass
                try:
                    self.edit_btn.configure(state='disabled')
                except Exception:
                    pass
                try:
                    self.delete_btn.configure(state='disabled')
                except Exception:
                    pass
                try:
                    self.save_btn.configure(state='disabled')
                except Exception:
                    pass
                try:
                    self.complete_btn.configure(state='disabled')
                except Exception:
                    pass
            if not can_export:
                try:
                    self.export_btn.configure(state='disabled')
                except Exception:
                    try:
                        self.export_btn.configure(fg_color='gray')
                    except Exception:
                        pass
        except Exception:
            pass
        self.refresh_btn = ctk.CTkButton(
            toolbar,
            text="Refresh",
            command=self._refresh_view,
            font=btn_font,
            height=34,
            fg_color="#334155",
            hover_color="#475569",
        )
        self.refresh_btn.grid(row=0, column=7, padx=4, pady=6, sticky="e")

        filter_bar = ctk.CTkFrame(left, fg_color="transparent")
        filter_bar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=(10, 8))
        filter_bar.grid_columnconfigure(1, weight=1)
        filter_bar.grid_columnconfigure(3, weight=1)
        ctk.CTkLabel(
            filter_bar,
            text="Status View",
            font=("Segoe UI Semibold", 13),
            text_color=self._text_primary,
        ).grid(row=0, column=0, sticky="w", padx=(0, 12))
        self._status_filter = ctk.StringVar(value="All")
        self._status_filter_control = ctk.CTkSegmentedButton(
            filter_bar,
            values=["All", "Normal", "Maintenance", "Due", "Overdue", "Critical"],
            command=lambda _value: self._refresh_view(),
            font=("Segoe UI Semibold", 12),
            selected_color=self._accent,
            unselected_color="#1f2937",
            unselected_hover_color="#334155",
        )
        self._status_filter_control.grid(row=0, column=1, sticky="ew")
        try:
            self._status_filter_control.set("All")
        except Exception:
            pass
        ctk.CTkLabel(
            filter_bar,
            text="Search",
            font=("Segoe UI Semibold", 13),
            text_color=self._text_primary,
        ).grid(row=1, column=0, sticky="w", padx=(0, 12), pady=(10, 0))
        self._search_var = ctk.StringVar(value="")
        self._search_entry = ctk.CTkEntry(
            filter_bar,
            textvariable=self._search_var,
            placeholder_text="Search by machine ID or name",
            font=("Segoe UI", 13),
            height=36,
        )
        self._search_entry.grid(row=1, column=1, columnspan=2, sticky="ew", pady=(10, 0))
        self._search_entry.bind("<KeyRelease>", lambda _event: self._refresh_view())
        clear_search_btn = ctk.CTkButton(
            filter_bar,
            text="Clear",
            width=72,
            height=36,
            command=self._clear_search,
            font=("Segoe UI Semibold", 12),
            fg_color="#334155",
            hover_color="#475569",
        )
        clear_search_btn.grid(row=1, column=3, sticky="e", pady=(10, 0))
        self._view_count_label = ctk.CTkLabel(
            filter_bar,
            text="0 machines",
            font=("Segoe UI", 12),
            text_color=self._text_muted,
        )
        self._view_count_label.grid(row=0, column=3, sticky="e", padx=(12, 0))

        # Treeview
        cols = [
            ("id", "ID"),
            ("registration_number", "Reg No"),
            ("company", "Company"),
            ("model", "Model"),
            ("maintenance_status", "Maint Status"),
            ("next_maintenance", "Next Maint"),
        ]
        self._tree = ttk.Treeview(left, columns=[c[0] for c in cols], show="headings", selectmode="browse", style="Machines.Treeview")
        try:
            style = ttk.Style()
            style.configure(
                "Machines.Treeview",
                font=("Segoe UI", 12),
                rowheight=36,
                background=self._surface_alt,
                fieldbackground=self._surface_alt,
                foreground=self._text_primary,
                borderwidth=0,
                relief="flat",
            )
            style.configure(
                "Machines.Treeview.Heading",
                font=("Segoe UI Semibold", 12),
                background=self._surface_header,
                foreground=self._text_primary,
                relief="flat",
            )
            style.map(
                "Machines.Treeview",
                background=[("selected", self._accent)],
                foreground=[("selected", "#ffffff")],
            )
            style.map(
                "Machines.Treeview.Heading",
                background=[("active", self._surface_header)],
                foreground=[("active", self._text_primary)],
            )
        except Exception:
            pass
        try:
            self._tree.tag_configure("even", background=self._surface_alt, foreground=self._text_primary)
            self._tree.tag_configure("odd", background="#111827", foreground=self._text_primary)
        except Exception:
            pass
        for key, heading in cols:
            self._tree.heading(key, text=heading)
            self._tree.column(key, width=160, anchor="w")

        vsb = ttk.Scrollbar(left, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(left, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)

        self._tree.bind("<<TreeviewSelect>>", self._on_select)
        self._tree.bind("<Double-1>", self._on_double)

        # Details
        self._detail_title = ctk.CTkLabel(
            right,
            text="Machine Details",
            font=("Segoe UI Semibold", 20),
            text_color=self._text_primary,
        )
        self._detail_title.pack(anchor="w", pady=(4,8))
        self._detail_badge = ctk.CTkLabel(
            right,
            text="Status: -",
            font=("Segoe UI Semibold", 12),
            text_color="#ffffff",
            fg_color="#334155",
            corner_radius=8,
            padx=10,
            pady=4,
        )
        self._detail_badge.pack(anchor="w", pady=(0, 8))
        self._detail_text = ctk.CTkLabel(
            right,
            text="Select a machine to see details",
            anchor="w",
            justify="left",
            font=("Segoe UI", 13),
            text_color=self._text_muted,
            wraplength=340,
        )
        self._detail_text.pack(fill="x")
        self._history_title = ctk.CTkLabel(
            right,
            text="Maintenance History",
            font=("Segoe UI Semibold", 16),
            text_color=self._text_primary,
        )
        self._history_title.pack(anchor="w", pady=(14, 6))
        self._history_box = ctk.CTkTextbox(
            right,
            height=220,
            corner_radius=10,
            fg_color=self._surface_alt,
            text_color=self._text_primary,
            font=("Segoe UI", 12),
            border_width=1,
            border_color="#1f2937",
        )
        self._history_box.pack(fill="both", expand=True)
        self._set_history_text("No maintenance history recorded yet.")

        self._refresh_view()

    # -- Data layer
    def _load_machines(self) -> None:
        """Load machines from the shared user-managed store."""
        try:
            rows = load_machines(include_archived=True)
            if rows:
                self.df = pd.DataFrame(rows)
            else:
                self.df = pd.DataFrame(columns=DEFAULT_COLUMNS)
            for col in DEFAULT_COLUMNS:
                if col not in self.df.columns:
                    self.df[col] = pd.NA
            self.df = self.df[DEFAULT_COLUMNS].copy()
            LOG.info("Loaded %d machines from user machine store", len(self.df))
        except Exception as exc:
            LOG.exception("Failed loading machine store: %s", exc)
            self.df = pd.DataFrame(columns=DEFAULT_COLUMNS)

    def _save_machines(self) -> None:
        try:
            # Only allow save for admin or maintainer
            try:
                from authz import has_role
                user = getattr(self, 'dashboard', None) and getattr(self.dashboard, 'user', None)
                if not user or not (has_role(user, 'admin') or has_role(user, 'maintainer')):
                    messagebox.showerror('Permission denied', 'You do not have permission to save machines.')
                    return
            except Exception:
                # if role check cannot run, deny to be safe
                messagebox.showerror('Permission denied', 'Unable to verify permissions.')
                return
            rows = []
            for record in self.df.to_dict("records"):
                cleaned = {}
                for key, value in record.items():
                    if isinstance(value, (list, dict)):
                        cleaned[key] = value
                        continue
                    try:
                        cleaned[key] = None if pd.isna(value) else value
                    except Exception:
                        cleaned[key] = value
                rows.append(cleaned)
            save_machines(rows)
            LOG.info("Saved %d machines to user machine store", len(rows))
            # notify dashboard if present
            if getattr(self, 'dashboard', None):
                try:
                    if hasattr(self.dashboard, '_data_cache'):
                        self.dashboard._data_cache.pop('machines', None)
                    if hasattr(self.dashboard, 'refresh_ui'):
                        self.dashboard.refresh_ui()
                except Exception:
                    LOG.exception("Failed to notify dashboard")
        except Exception as exc:
            LOG.exception("Failed to save machines file: %s", exc)
            messagebox.showerror("Save error", f"Failed to save machines: {exc}")

    def _export_csv(self) -> None:
        """Export current machines view to a CSV file in data/exports."""
        try:
            # restrict exports to admin or reporter role
            try:
                from authz import has_role
                user = getattr(self, 'dashboard', None) and getattr(self.dashboard, 'user', None)
                if not user or not (has_role(user, 'admin') or has_role(user, 'reporter')):
                    messagebox.showerror('Permission denied', 'You do not have permission to export machines.')
                    return
            except Exception:
                messagebox.showerror('Permission denied', 'Unable to verify permissions.')
                return
            export_dir = str(exports_dir())
            os.makedirs(export_dir, exist_ok=True)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            fname_csv = os.path.join(export_dir, f"machines_export_{ts}.csv")
            fname_xlsx = os.path.join(export_dir, f"machines_export_{ts}.xlsx")
            # prefer unarchived view
            if self.df.empty:
                df = self.df
            else:
                df = self.df[~(self.df.get('archived') == True)]
            # write CSV
            df.to_csv(fname_csv, index=False)
            # write Excel (pandas -> openpyxl) then style if possible
            try:
                df.to_excel(fname_xlsx, index=False)
                if load_workbook and Font is not None:
                    wb = load_workbook(fname_xlsx)
                    ws = wb.active
                    # header styling
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="4F81BD")
                    for cell in list(ws[1]):
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # thin border
                    thin = Side(border_style="thin", color="CCCCCC")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # alternating row fill and borders
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        rnum = row[0].row
                        if rnum % 2 == 0:
                            fill = PatternFill("solid", fgColor="F7F7F7")
                        else:
                            fill = PatternFill("solid", fgColor="FFFFFF")
                        for cell in row:
                            cell.border = border
                            cell.fill = fill
                            cell.alignment = Alignment(vertical='top')
                    # auto-size columns
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                val = str(cell.value) if cell.value is not None else ""
                            except Exception:
                                val = ""
                            if len(val) > max_length:
                                max_length = len(val)
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[col_letter].width = adjusted_width
                    wb.save(fname_xlsx)
            except Exception:
                LOG.exception("Failed to write styled Excel; CSV saved")

            messagebox.showinfo("Exported", f"Exported machines to:\n{fname_csv}\n{fname_xlsx}")
        except Exception as exc:
            LOG.exception("Failed to export machines: %s", exc)
            messagebox.showerror("Export error", f"Failed to export machines: {exc}")

    # -- UI handlers
    def _refresh_view(self) -> None:
        try:
            # clear tree
            for iid in self._tree.get_children():
                self._tree.delete(iid)
            # build display df (exclude archived)
            if self.df.empty:
                display_df = self.df
            else:
                display_df = self.df[~(self.df.get('archived') == True)]
            display_df = self._apply_status_filter(display_df)
            display_df = self._apply_search_filter(display_df)

            for rownum, (idx, row) in enumerate(display_df.iterrows()):
                status_info = self._status_visuals(row.to_dict())
                next_window = row.get('next_maintenance') or row.get('due_date') or row.get('next_due_hours') or ''
                vals = [
                    row.get('id', ''),
                    row.get('registration_number', ''),
                    row.get('company', ''),
                    row.get('model', ''),
                    status_info["label"],
                    next_window,
                ]
                tag = 'even' if (rownum % 2 == 0) else 'odd'
                self._tree.insert('', 'end', iid=str(idx), values=vals, tags=(tag,))
            try:
                total = len(display_df.index)
                self._view_count_label.configure(text=f"{total} machine{'s' if total != 1 else ''}")
            except Exception:
                pass

            # clear details and hide details pane so table fills full width
            self._detail_text.configure(text='Select a machine to see details')
            try:
                self._detail_badge.configure(text="Status: -", fg_color="#334155")
            except Exception:
                pass
            self._set_history_text("No maintenance history recorded yet.")
            try:
                if getattr(self, '_right_visible', False):
                    self._hide_details()
            except Exception:
                pass
        except Exception as exc:
            LOG.exception("Error refreshing machines view: %s", exc)
            # visible banner
            banner = ctk.CTkLabel(self, text=f"Error loading machines: {exc}", fg_color="#b00020", text_color="#ffffff")
            banner.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=8)

    def _on_select(self, ev=None) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except Exception:
            return
        try:
            r = self.df.loc[idx].to_dict()
        except Exception:
            return
        notes_val = r.get('notes', '')
        try:
            if pd.isna(notes_val):
                notes_val = ''
        except Exception:
            pass
        notes_str = str(notes_val)[:200]
        status_info = self._status_visuals(r)
        lines = [
            f"ID: {r.get('id','')}",
            f"Name: {r.get('name','')}",
            f"Model: {r.get('model','')}",
            f"Company: {r.get('company','')}",
            f"Status: {status_info['label']}",
            f"Running hours: {r.get('current_hours') or r.get('hours') or ''}",
            f"Next due hours: {r.get('next_due_hours','')}",
            f"Next maintenance: {r.get('next_maintenance','')}",
            f"Notes: {notes_str}",
        ]
        txt = "\n\n".join(lines)
        # ensure details pane is visible and update text
        try:
            self._show_details()
        except Exception:
            pass
        try:
            self._detail_badge.configure(text=status_info["label"], fg_color=status_info["color"])
        except Exception:
            pass
        self._detail_text.configure(text=txt)
        self._set_history_text(self._format_machine_history(r))

    def _status_visuals(self, row: dict) -> dict:
        try:
            status = str(evaluate_machine_status(row).get("status") or "normal").lower()
        except Exception:
            status = str(row.get("status") or "normal").lower()
        palette = {
            "normal": {"label": "Normal", "color": "#0f766e"},
            "maintenance": {"label": "Maintenance", "color": "#d97706"},
            "due": {"label": "Due", "color": "#dc2626"},
            "overdue": {"label": "Overdue", "color": "#991b1b"},
            "critical": {"label": "Critical", "color": "#7f1d1d"},
        }
        return palette.get(status, {"label": status.title() or "Normal", "color": "#334155"})

    def _apply_status_filter(self, display_df: pd.DataFrame) -> pd.DataFrame:
        try:
            selected = str(self._status_filter_control.get() or "All").strip().lower()
        except Exception:
            selected = "all"
        if selected == "all" or display_df.empty:
            return display_df

        rows = []
        for idx, row in display_df.iterrows():
            status = str(evaluate_machine_status(row.to_dict()).get("status") or "normal").strip().lower()
            if status == selected:
                rows.append(idx)
        if not rows:
            return display_df.iloc[0:0]
        return display_df.loc[rows]

    def _apply_search_filter(self, display_df: pd.DataFrame) -> pd.DataFrame:
        try:
            query = str(self._search_var.get() or "").strip().lower()
        except Exception:
            query = ""
        if not query or display_df.empty:
            return display_df

        rows = []
        for idx, row in display_df.iterrows():
            haystacks = [
                str(row.get("id") or "").lower(),
                str(row.get("name") or "").lower(),
                str(row.get("model") or "").lower(),
            ]
            if any(query in item for item in haystacks if item):
                rows.append(idx)
        if not rows:
            return display_df.iloc[0:0]
        return display_df.loc[rows]

    def _clear_search(self) -> None:
        try:
            self._search_var.set("")
        except Exception:
            pass
        self._refresh_view()

    def _set_history_text(self, text: str) -> None:
        try:
            self._history_box.configure(state="normal")
            self._history_box.delete("1.0", "end")
            self._history_box.insert("1.0", text)
            self._history_box.configure(state="disabled")
        except Exception:
            pass

    def _format_machine_history(self, row: dict) -> str:
        history = machine_history(row)
        if not history:
            return "No maintenance completion history recorded yet.\n\nUse `Complete Maint` after servicing a machine to build the history log."
        chunks = []
        for item in history[:8]:
            completed_at = str(item.get("completed_at") or "-").replace("T", " ")
            previous_status = str(item.get("previous_status") or "normal").title()
            current_hours = item.get("current_hours")
            rolled_hours = item.get("rolled_next_due_hours")
            rolled_date = item.get("rolled_due_date") or "-"
            completed_by = item.get("completed_by") or "System"
            notes = str(item.get("completion_notes") or "").strip()
            lines = [
                f"{completed_at}",
                f"Completed by: {completed_by}",
                f"Previous status: {previous_status}",
                f"Runtime at service: {current_hours if current_hours not in (None, '') else '-'} hrs",
                f"Rolled next due date: {rolled_date}",
                f"Rolled next due hours: {rolled_hours if rolled_hours not in (None, '') else '-'}",
            ]
            if notes:
                lines.append(f"Notes: {notes}")
            chunks.append("\n".join(lines))
        return "\n\n".join(chunks)

    def _show_details(self) -> None:
        """Show the right-hand details pane and let the table occupy remaining width."""
        if getattr(self, '_right_visible', False):
            return
        try:
            self._right.grid()
            self._right_visible = True
            # make sure left column remains expandable
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=0)
        except Exception:
            pass

    def _hide_details(self) -> None:
        """Hide the right-hand details pane so the table fills the full width."""
        if not getattr(self, '_right_visible', False):
            return
        try:
            self._right.grid_remove()
            self._right_visible = False
            # ensure column 0 expands to full width
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=0)
        except Exception:
            pass

    def _on_double(self, ev=None) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        try:
            idx = int(sel[0])
        except Exception:
            return
        self._open_edit_dialog(idx)

    def _edit_selected(self) -> None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Edit", "Please select a machine to edit")
            return
        try:
            idx = int(sel[0])
        except Exception:
            messagebox.showerror("Edit", "Invalid selection")
            return
        self._open_edit_dialog(idx)

    def _delete_selected(self) -> None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Delete", "Please select a machine to delete")
            return
        try:
            idx = int(sel[0])
        except Exception:
            messagebox.showerror("Delete", "Invalid selection")
            return
        try:
            row = self.df.loc[idx].to_dict()
        except Exception:
            return
        # Role check: only admin may delete/archive machines
        try:
            from authz import has_role
            user = getattr(self, 'dashboard', None) and getattr(self.dashboard, 'user', None)
            if not user or not has_role(user, 'admin'):
                messagebox.showerror("Permission denied", "You must be an administrator to archive machines.")
                return
        except Exception:
            # if role check fails for any reason, deny delete to be safe
            try:
                messagebox.showerror("Permission denied", "Unable to verify permissions. Contact your administrator.")
            except Exception:
                pass
            return

        if messagebox.askyesno("Delete", f"Archive machine '{row.get('name','')}'?"):
            try:
                archive_machine(str(row.get("id") or ""))
            except Exception:
                self.df.at[idx, 'archived'] = True
                self.df.at[idx, 'last_updated'] = datetime.datetime.now().isoformat(sep=' ', timespec='seconds')
                self._save_machines()
            else:
                self._load_machines()
            self._refresh_view()

    def select_machine(self, machine_id: str) -> None:
        target = str(machine_id or "").strip()
        if not target:
            return
        try:
            self._status_filter_control.set("All")
        except Exception:
            pass
        try:
            self._search_var.set(target)
        except Exception:
            pass
        self._refresh_view()
        for iid in self._tree.get_children():
            values = self._tree.item(iid, "values") or []
            if values and str(values[0]).strip() == target:
                self._tree.selection_set(iid)
                self._tree.focus(iid)
                self._tree.see(iid)
                self._on_select()
                break

    def _complete_selected_maintenance(self) -> None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Complete Maintenance", "Please select a machine first")
            return
        try:
            user = getattr(self, 'dashboard', None) and getattr(self.dashboard, 'user', None)
            if not user or not (has_role(user, 'admin') or has_role(user, 'maintainer')):
                messagebox.showerror("Permission denied", "Only admin or maintainer can complete maintenance.")
                return
        except Exception:
            messagebox.showerror("Permission denied", "Unable to verify permissions.")
            return
        try:
            idx = int(sel[0])
            current = self.df.loc[idx].to_dict()
        except Exception:
            messagebox.showerror("Complete Maintenance", "Invalid selection")
            return

        if not messagebox.askyesno(
            "Complete Maintenance",
            f"Mark maintenance completed for '{current.get('name','')}' and roll the next due date/hours forward?",
        ):
            return

        try:
            completed_by = ""
            dashboard_user = getattr(self, "dashboard", None) and getattr(self.dashboard, "user", None)
            if isinstance(dashboard_user, dict):
                completed_by = str(
                    dashboard_user.get("display_name")
                    or dashboard_user.get("username")
                    or dashboard_user.get("email")
                    or ""
                ).strip()
            updated = complete_machine_maintenance(current, completed_by=completed_by or None)
            for key, value in updated.items():
                if key not in self.df.columns:
                    self.df[key] = pd.NA
                self.df.at[idx, key] = value
            self._save_machines()
            self._refresh_view()

            completion_sms_sent = 0
            completion_sms_failures = 0
            try:
                try:
                    from ..machine_alert_runner import send_maintenance_completion_sms
                    from ..settings_store import load_settings
                except Exception:
                    from machine_alert_runner import send_maintenance_completion_sms
                    from settings_store import load_settings
                sms_result = send_maintenance_completion_sms(
                    updated,
                    completed_by=completed_by or "",
                    settings=load_settings(),
                )
                completion_sms_sent = int(sms_result.get("sent", 0) or 0)
                completion_sms_failures = int(sms_result.get("failures", 0) or 0)
            except Exception:
                completion_sms_sent = 0
                completion_sms_failures = 0

            messagebox.showinfo(
                "Maintenance Completed",
                f"Next maintenance date: {updated.get('next_maintenance') or '-'}\n"
                f"Next due hours: {updated.get('next_due_hours') or '-'}\n"
                f"Completion SMS sent: {completion_sms_sent} (failed: {completion_sms_failures})",
            )
        except Exception as exc:
            LOG.exception("Failed to complete maintenance: %s", exc)
            messagebox.showerror("Complete Maintenance", f"Failed to complete maintenance: {exc}")

    def _print_completion_report_selected(self) -> None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Completion Report", "Please select a machine first")
            return
        try:
            idx = int(sel[0])
            machine = self.df.loc[idx].to_dict()
        except Exception:
            messagebox.showerror("Completion Report", "Invalid selection")
            return

        history_rows = machine_history(machine)
        latest = history_rows[0] if history_rows else {}
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_id = str(machine.get("id") or "machine").replace(" ", "_")
        output_path = exports_dir() / f"maintenance_completion_{safe_id}_{timestamp}.html"
        try:
            output_path.write_text(self._build_completion_report_html(machine, latest), encoding="utf-8")
            messagebox.showinfo(
                "Completion Report Ready",
                f"Printable maintenance completion report saved to:\n{output_path}",
            )
        except Exception as exc:
            LOG.exception("Failed to generate completion report: %s", exc)
            messagebox.showerror("Completion Report", f"Failed to create report: {exc}")

    def _build_completion_report_html(self, machine: dict, latest: dict) -> str:
        machine_label = escape(str(machine.get("name") or machine.get("model") or machine.get("id") or "Machine"))
        machine_id = escape(str(machine.get("id") or "-"))
        status_info = self._status_visuals(machine)
        try:
            mine = get_active_mine() or {}
        except Exception:
            mine = {}
        mine_name = escape(str(mine.get("mine_name") or "Unassigned Mine"))
        mine_company = escape(str(mine.get("company_name") or "Not set"))
        mine_quarry_type = escape(str(mine.get("quarry_type") or "Not set"))
        mine_lease_area = escape(str(mine.get("lease_area") or "Not set"))
        mine_address = escape(str(mine.get("address") or "Not set"))
        mine_maps = escape(str(mine.get("google_maps_link") or "Not set"))
        completed_at = escape(str(latest.get("completed_at") or machine.get("last_maintenance_completed_at") or "-").replace("T", " "))
        completed_by = escape(str(latest.get("completed_by") or "System"))
        previous_status = escape(str(latest.get("previous_status") or machine.get("last_maintenance_status") or "-").title())
        rolled_date = escape(str(latest.get("rolled_due_date") or machine.get("next_maintenance") or machine.get("due_date") or "-"))
        rolled_hours = escape(str(latest.get("rolled_next_due_hours") or machine.get("next_due_hours") or "-"))
        current_hours = escape(str(latest.get("current_hours") or machine.get("current_hours") or machine.get("hours") or "-"))
        notes = escape(str(latest.get("completion_notes") or machine.get("notes") or "No completion notes recorded."))
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Maintenance Completion Report - {machine_id}</title>
  <style>
    :root {{
      --ink: #0f172a;
      --muted: #475569;
      --line: #cbd5e1;
      --panel: #f8fafc;
      --accent: #2563eb;
      --badge: {status_info["color"]};
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: #e2e8f0; color: var(--ink); font-family: "Segoe UI", Arial, sans-serif; }}
    .page {{ max-width: 980px; margin: 28px auto; background: white; border-radius: 18px; overflow: hidden; box-shadow: 0 24px 60px rgba(15, 23, 42, 0.14); }}
    .hero {{ padding: 28px 34px; background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 100%); color: white; }}
    .hero h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .hero p {{ margin: 0; font-size: 14px; color: rgba(255,255,255,0.82); }}
    .badge {{ display: inline-block; margin-top: 14px; padding: 7px 12px; border-radius: 999px; background: var(--badge); font-weight: 700; font-size: 12px; letter-spacing: 0.04em; text-transform: uppercase; }}
    .site-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 14px; }}
    .section {{ padding: 24px 34px; border-top: 1px solid var(--line); }}
    .section h2 {{ margin: 0 0 14px; font-size: 18px; }}
    .grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 14px; }}
    .stat {{ background: var(--panel); border: 1px solid var(--line); border-radius: 14px; padding: 14px 16px; }}
    .label {{ font-size: 12px; text-transform: uppercase; letter-spacing: 0.06em; color: var(--muted); margin-bottom: 6px; }}
    .value {{ font-size: 17px; font-weight: 700; overflow-wrap: anywhere; word-break: break-word; }}
    .notes {{ white-space: pre-wrap; line-height: 1.6; font-size: 14px; overflow-wrap: anywhere; word-break: break-word; }}
    .footer {{ padding: 16px 34px 26px; color: var(--muted); font-size: 12px; }}
    @media print {{
      body {{ background: white; }}
      .page {{ margin: 0; box-shadow: none; border-radius: 0; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <section class="hero">
      <h1>Maintenance Completion Report</h1>
      <p>Mining Maintenance System | {mine_name} | Generated {escape(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</p>
      <div class="badge">{escape(status_info["label"])}</div>
    </section>
    <section class="section">
      <h2>Mine Site Context</h2>
      <div class="site-grid">
        <div class="stat"><div class="label">Active Mine</div><div class="value">{mine_name}</div></div>
        <div class="stat"><div class="label">Company</div><div class="value">{mine_company}</div></div>
        <div class="stat"><div class="label">Quarry Type</div><div class="value">{mine_quarry_type}</div></div>
        <div class="stat"><div class="label">Lease Area</div><div class="value">{mine_lease_area}</div></div>
        <div class="stat"><div class="label">Address</div><div class="value">{mine_address}</div></div>
        <div class="stat"><div class="label">Google Maps</div><div class="value">{mine_maps}</div></div>
      </div>
    </section>
    <section class="section">
      <h2>Machine Profile</h2>
      <div class="grid">
        <div class="stat"><div class="label">Machine ID</div><div class="value">{machine_id}</div></div>
        <div class="stat"><div class="label">Machine Name</div><div class="value">{machine_label}</div></div>
        <div class="stat"><div class="label">Model</div><div class="value">{escape(str(machine.get("model") or "-"))}</div></div>
        <div class="stat"><div class="label">Company</div><div class="value">{escape(str(machine.get("company") or "-"))}</div></div>
      </div>
    </section>
    <section class="section">
      <h2>Completion Snapshot</h2>
      <div class="grid">
        <div class="stat"><div class="label">Completed At</div><div class="value">{completed_at}</div></div>
        <div class="stat"><div class="label">Completed By</div><div class="value">{completed_by}</div></div>
        <div class="stat"><div class="label">Previous Status</div><div class="value">{previous_status}</div></div>
        <div class="stat"><div class="label">Runtime At Service</div><div class="value">{current_hours} hrs</div></div>
        <div class="stat"><div class="label">Next Due Date</div><div class="value">{rolled_date}</div></div>
        <div class="stat"><div class="label">Next Due Hours</div><div class="value">{rolled_hours}</div></div>
      </div>
    </section>
    <section class="section">
      <h2>Service Notes</h2>
      <div class="notes">{notes}</div>
    </section>
    <div class="footer">This printable report is generated directly from the app's saved machine maintenance history.</div>
  </div>
</body>
</html>
"""

    def _open_add_dialog(self) -> None:
        # provide a thin shim - _open_machine_dialog may be missing in older code
        try:
            self._open_machine_dialog()
        except Exception:
            self._open_edit_dialog(None)

    def _open_machine_dialog(self) -> None:
        """Compatibility shim that opens the add-machine dialog."""
        self._open_edit_dialog(None)

    def _open_edit_dialog(self, index: Optional[int] = None) -> None:
        initial = None
        if index is not None:
            try:
                initial = self.df.loc[index].to_dict()
            except Exception:
                initial = None
        dlg, dlg_destroy = create_dialog(self, title=("Edit Machine" if initial else "Add Machine"), width=560, height=520)

        labels = [
            ("Machine ID", "id"),
            ("Name", "name"),
            ("Type", "type"),
            ("Registration", "registration_number"),
            ("Company", "company"),
            ("Model", "model"),
            ("Operator Phone", "operator_phone"),
            ("Running Hours", "hours"),
            ("Service Interval Hours", "service_interval_hours"),
            ("Next Due Hours", "next_due_hours"),
            ("Alert Before Hours", "hour_alert_window"),
            ("Overdue After Hours", "hour_overdue_after_hours"),
            ("Service Date (YYYY-MM-DD)", "service_date"),
            ("Due Date (YYYY-MM-DD)", "due_date"),
            ("Next maintenance (YYYY-MM-DD)", "next_maintenance"),
            ("Status", "status"),
        ]
        vars_map = {}
        for i, (lbl, key) in enumerate(labels):
            ctk.CTkLabel(dlg, text=lbl, font=("Segoe UI", 13)).grid(row=i, column=0, padx=12, pady=6, sticky='w')
            v = ctk.StringVar(value=(str(initial.get(key)) if initial and initial.get(key) is not None else ""))
            entry = ctk.CTkEntry(dlg, textvariable=v, font=("Segoe UI", 13), height=34)
            entry.grid(row=i, column=1, padx=12, pady=6, sticky='ew')
            vars_map[key] = v
        dlg.grid_columnconfigure(1, weight=1)

        def on_save():
            mid = vars_map['id'].get().strip()
            name = vars_map['name'].get().strip()
            reg = vars_map['registration_number'].get().strip()
            comp = vars_map['company'].get().strip()
            model = vars_map['model'].get().strip()
            mtype = vars_map['type'].get().strip() or "Machine"
            operator_phone = vars_map['operator_phone'].get().strip()
            hours = vars_map['hours'].get().strip()
            service_interval_hours = vars_map['service_interval_hours'].get().strip()
            next_due_hours = vars_map['next_due_hours'].get().strip()
            hour_alert_window = vars_map['hour_alert_window'].get().strip()
            hour_overdue_after_hours = vars_map['hour_overdue_after_hours'].get().strip()
            service_date = vars_map['service_date'].get().strip()
            due_date = vars_map['due_date'].get().strip()
            next_maint = vars_map['next_maintenance'].get().strip()
            status = vars_map['status'].get().strip().lower() or "normal"
            if not validate_required(mid, "Machine ID"):
                return
            if not validate_required(name, "Name"):
                return
            if not validate_optional_phone(operator_phone, "Operator Phone"):
                return
            if not validate_number(hours, "Running Hours", minimum=0):
                return
            if not validate_number(service_interval_hours, "Service Interval Hours", minimum=0):
                return
            if not validate_number(next_due_hours, "Next Due Hours", minimum=0):
                return
            if not validate_number(hour_alert_window, "Alert Before Hours", minimum=0):
                return
            if not validate_number(hour_overdue_after_hours, "Overdue After Hours", minimum=0):
                return
            if not validate_date_string(service_date, "Service Date"):
                return
            if not validate_date_string(due_date, "Due Date"):
                return
            if not validate_date_string(next_maint, "Next Maintenance"):
                return

            try:
                existing_ids = {
                    str(value).strip()
                    for value in self.df.get('id', pd.Series(dtype=str)).tolist()
                    if str(value).strip()
                }
                current_id = str(initial.get("id") or "").strip() if initial else ""
                if current_id:
                    existing_ids.discard(current_id)
                if mid in existing_ids:
                    messagebox.showerror("Validation", f"Machine ID '{mid}' already exists")
                    return
            except Exception:
                pass

            normalized_phone = normalize_phone_input(operator_phone) or pd.NA
            normalized_service_date = normalize_date_input(service_date) or pd.NA
            normalized_due_date = normalize_date_input(due_date) or pd.NA
            normalized_next_maint = normalize_date_input(next_maint) or pd.NA
            normalized_hours = str(float(hours)).rstrip("0").rstrip(".") if hours else pd.NA
            normalized_service_interval = str(float(service_interval_hours)).rstrip("0").rstrip(".") if service_interval_hours else pd.NA
            normalized_next_due_hours = str(float(next_due_hours)).rstrip("0").rstrip(".") if next_due_hours else pd.NA
            normalized_alert_window = str(float(hour_alert_window)).rstrip("0").rstrip(".") if hour_alert_window else pd.NA
            normalized_overdue_after_hours = str(float(hour_overdue_after_hours)).rstrip("0").rstrip(".") if hour_overdue_after_hours else pd.NA

            if normalized_next_due_hours is pd.NA and normalized_service_interval is not pd.NA and normalized_hours is not pd.NA:
                try:
                    normalized_next_due_hours = str(
                        round(float(normalized_hours) + float(normalized_service_interval), 2)
                    ).rstrip("0").rstrip(".")
                except Exception:
                    pass

            if normalized_due_date is not pd.NA and normalized_service_date is not pd.NA:
                try:
                    if normalized_due_date < normalized_service_date:
                        messagebox.showerror("Validation", "Due Date cannot be before Service Date")
                        return
                except Exception:
                    pass
            if normalized_next_maint is not pd.NA and normalized_service_date is not pd.NA:
                try:
                    if normalized_next_maint < normalized_service_date:
                        messagebox.showerror("Validation", "Next Maintenance cannot be before Service Date")
                        return
                except Exception:
                    pass

            if status not in {"normal", "maintenance", "due", "overdue", "critical"}:
                messagebox.showerror("Validation", "Status must be normal, maintenance, due, overdue, or critical")
                return
            ts = datetime.datetime.now().isoformat(sep=' ', timespec='seconds')
            if initial is None:
                # create
                new = {col: pd.NA for col in DEFAULT_COLUMNS}
                new.update({
                    'id': mid or '',
                    'name': name,
                    'type': mtype,
                    'registration_number': reg,
                    'company': comp,
                    'model': model,
                    'operator_phone': normalized_phone,
                    'hours': normalized_hours,
                    'current_hours': normalized_hours,
                    'service_interval_hours': normalized_service_interval,
                    'next_due_hours': normalized_next_due_hours,
                    'hour_alert_window': normalized_alert_window,
                    'hour_overdue_after_hours': normalized_overdue_after_hours,
                    'service_date': normalized_service_date,
                    'due_date': normalized_due_date,
                    'next_maintenance': normalized_next_maint,
                    'status': status,
                    'maintenance_status': status,
                    'created_at': ts,
                    'last_updated': ts,
                    'archived': False,
                })
                self.df = pd.concat([self.df, pd.DataFrame([new])], ignore_index=True)
            else:
                # update
                try:
                    if mid:
                        self.df.at[index, 'id'] = mid
                    self.df.at[index, 'name'] = name
                    self.df.at[index, 'type'] = mtype
                    self.df.at[index, 'registration_number'] = reg
                    self.df.at[index, 'company'] = comp
                    self.df.at[index, 'model'] = model
                    self.df.at[index, 'operator_phone'] = normalized_phone
                    self.df.at[index, 'hours'] = normalized_hours
                    self.df.at[index, 'current_hours'] = normalized_hours
                    self.df.at[index, 'service_interval_hours'] = normalized_service_interval
                    self.df.at[index, 'next_due_hours'] = normalized_next_due_hours
                    self.df.at[index, 'hour_alert_window'] = normalized_alert_window
                    self.df.at[index, 'hour_overdue_after_hours'] = normalized_overdue_after_hours
                    self.df.at[index, 'service_date'] = normalized_service_date
                    self.df.at[index, 'due_date'] = normalized_due_date
                    self.df.at[index, 'next_maintenance'] = normalized_next_maint
                    self.df.at[index, 'status'] = status
                    self.df.at[index, 'maintenance_status'] = status
                    self.df.at[index, 'last_updated'] = ts
                except Exception:
                    LOG.exception("Failed to update machine %s", index)
            try:
                self._status_filter_control.set("All")
            except Exception:
                pass
            try:
                self._search_var.set("")
            except Exception:
                pass
            self._save_machines()
            try:
                dlg_destroy()
            except Exception:
                try:
                    dlg.destroy()
                except Exception:
                    pass
            self._refresh_view()

        btn_frame = ctk.CTkFrame(dlg, fg_color='transparent')
        btn_frame.grid(row=len(labels), column=0, columnspan=2, sticky='ew', padx=12, pady=12)
        save_btn = ctk.CTkButton(
            btn_frame,
            text='Save',
            command=on_save,
            font=("Segoe UI Semibold", 13),
            fg_color="#059669",
            hover_color="#047857",
        )
        save_btn.pack(side='right', padx=6)
        cancel_btn = ctk.CTkButton(
            btn_frame,
            text='Cancel',
            command=dlg_destroy,
            font=("Segoe UI", 13),
            fg_color="#334155",
            hover_color="#475569",
        )
        cancel_btn.pack(side='right', padx=6)
